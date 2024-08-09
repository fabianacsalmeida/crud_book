import tkinter as tk
from tkinter import messagebox
import openpyxl


class Book:
    def __init__(self, title, author, year):
        self.title = title
        self.author = author
        self.year = year


class Database:
    def __init__(self, filename='books.xlsx'):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename) if self.file_exists() else openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.setup_sheet()

    def file_exists(self):
        try:
            open(self.filename, 'r').close()
            return True
        except FileNotFoundError:
            return False

    def setup_sheet(self):
        if not self.sheet['A1'].value:
            self.sheet.append(['Title', 'Author', 'Year'])
            self.workbook.save(self.filename)

    def add_book(self, book):
        self.sheet.append([book.title, book.author, book.year])
        self.workbook.save(self.filename)

    def get_books(self):
        books = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            books.append(Book(*row))
        return books

    def update_book(self, row, book):
        self.sheet[f'A{row}'] = book.title
        self.sheet[f'B{row}'] = book.author
        self.sheet[f'C{row}'] = book.year
        self.workbook.save(self.filename)

    def delete_book(self, row):
        self.sheet.delete_rows(row)
        self.workbook.save(self.filename)

import tkinter as tk
from tkinter import messagebox

class MainApp:
    def __init__(self, root, db):
        self.root = root
        self.db = db
        self.root.title("Gerenciador de Livros")

        self.book_listbox = tk.Listbox(root, width=50, height=15)
        self.book_listbox.pack()

        self.add_button = tk.Button(root, text="Adicionar Livro", command=self.open_add_book_window)
        self.add_button.pack()

        self.edit_button = tk.Button(root, text="Editar Livro", command=self.open_edit_book_window)
        self.edit_button.pack()

        self.delete_button = tk.Button(root, text="Remover Livro", command=self.delete_book)
        self.delete_button.pack()

        self.load_books()

    def load_books(self):
        self.book_listbox.delete(0, tk.END)
        books = self.db.get_books()
        for idx, book in enumerate(books, start=2):
            self.book_listbox.insert(tk.END, f"{idx - 1} - {book.title} - {book.author} - {book.year}")

    def open_add_book_window(self):
        self.new_window = tk.Toplevel(self.root)
        self.app = AddEditBookWindow(self.new_window, self.db, self)

    def open_edit_book_window(self):
        selected_book = self.book_listbox.curselection()
        if selected_book:
            row = selected_book[0] + 2
            self.new_window = tk.Toplevel(self.root)
            self.app = AddEditBookWindow(self.new_window, self.db, self, row)
        else:
            messagebox.showwarning("Seleção Inválida", "Por favor, selecione um livro para editar.")

    def delete_book(self):
        selected_book = self.book_listbox.curselection()
        if selected_book:
            row = selected_book[0] + 2
            self.db.delete_book(row)
            self.load_books()
        else:
            messagebox.showwarning("Seleção Inválida", "Por favor, selecione um livro para remover.")

class AddEditBookWindow:
    def __init__(self, root, db, main_app, row=None):
        self.root = root
        self.db = db
        self.main_app = main_app
        self.row = row
        self.root.title("Adicionar Livro" if row is None else "Editar Livro")

        self.label_title = tk.Label(root, text="Título:")
        self.label_title.pack()
        self.entry_title = tk.Entry(root)
        self.entry_title.pack()

        self.label_author = tk.Label(root, text="Autor:")
        self.label_author.pack()
        self.entry_author = tk.Entry(root)
        self.entry_author.pack()

        self.label_year = tk.Label(root, text="Ano:")
        self.label_year.pack()
        self.entry_year = tk.Entry(root)
        self.entry_year.pack()

        self.save_button = tk.Button(root, text="Salvar", command=self.save_book)
        self.save_button.pack()

        if row:
            self.load_book()

    def load_book(self):
        book = self.db.get_books()[self.row - 2]
        self.entry_title.insert(0, book.title)
        self.entry_author.insert(0, book.author)
        self.entry_year.insert(0, book.year)

    def save_book(self):
        title = self.entry_title.get()
        author = self.entry_author.get()
        year = self.entry_year.get()
        book = Book(title, author, year)

        if self.row:
            self.db.update_book(self.row, book)
        else:
            self.db.add_book(book)

        self.main_app.load_books()
        self.root.destroy()


if __name__ == "__main__":
    db = Database()
    root = tk.Tk()
    app = MainApp(root, db)
    root.mainloop()
